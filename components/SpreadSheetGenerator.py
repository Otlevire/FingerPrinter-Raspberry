import openpyxl
import datetime
from dao.ReportDAO import ReportDAO
class SpreadSheetGenerator:

    def __init__(self):
        self.days_of_week =['SEGUNDA', 'TERCA', 'QUARTA', 'QUINTA', 'SEXTA', 'SABADO', 'DOMINGO']
        self.months_of_year = ['JANEIRO', 'FEVEREIRO', 'MARÇO', 'ABRIL', 'MAIO', 'JUNHO', 'JULHO', 'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO']
        self.SPREAD_SHEET_FILENAME = f'PLANILHA-DE-{self.months_of_year[datetime.date.today().month-1]}-{datetime.date.today().year}.xlsx'

    def generate_spread_sheet(self, registers, filename=None):
        book = openpyxl.Workbook()
        for register in registers:
            if len(register) > 4:
                DAY_SHEET =f'{register[3]}-{register[4]}'
                if DAY_SHEET not in book.sheetnames:
                    book.create_sheet(DAY_SHEET)
                    page = book[DAY_SHEET]
                    page.append(['ID', 'Num Mecanografico', 'Tipo','DoW','Dia', 'Mês', 'Ano', 'Horario'])
                else:
                    page = book[DAY_SHEET]
                page.append(register)
        try:
            if filename==None:
                book.save(f'reports/{self.SPREAD_SHEET_FILENAME}')
            else:
                book.save(f'{filename}/{self.SPREAD_SHEET_FILENAME}')
            return True
        except Exception as e:
            return False

if __name__ == '__main__':

    spread_sheet_generator = SpreadSheetGenerator()
    spread_sheet_generator.generate_spread_sheet(7)




